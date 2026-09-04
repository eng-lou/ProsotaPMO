from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from fastapi import HTTPException

# P6's own "Export to Excel" TASK-sheet layout (2026-09-04, per Maro — a
# series of monthly progress extracts, used to layer real historical
# progress onto an already-PMXML-imported schedule). Confirmed directly
# against real files (P6_Extract_2011-05-01.xlsx through -10-01.xlsx): one
# "TASK" sheet, header row 1, one row per real Activity (no WBS/summary
# rows) — a much flatter report than a full PMXML export (no calendars,
# relationships, resource rates), but P6 has already computed real
# per-activity PV/AC/EV/BAC as of that extract's own data date, so this
# reads those numbers straight rather than re-deriving anything. Matched by
# HEADER NAME, not column position — the four cost columns carry a real "£"
# in their own header text (confirmed: U+00A3), matched by prefix here so
# this file's own source encoding never has to round-trip that character.
_HEADER_PREFIXES = {
    "activity_id": "Activity ID",
    "status": "Activity Status",
    "start": "(*)Start",
    "finish": "(*)Finish",
    "actual_cost": "(*)Actual Cost(",
    "earned_value_cost": "(*)Earned Value Cost(",
    "bac": "(*)Budget At Completion(",
}


@dataclass
class ParsedProgressRow:
    activity_id: str  # P6's own Activity ID, e.g. "EC1090" — matched against the "P6 Activity ID" UDF
    status: str  # "Not Started" | "In Progress" | "Completed" (P6's own text, passed through verbatim)
    start: datetime | None
    finish: datetime | None
    actual_cost: Decimal | None
    earned_value_cost: Decimal | None
    bac: Decimal | None


def _decimal(v: object) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return None


def parse_p6_excel_progress(data: bytes) -> list[ParsedProgressRow]:
    """Parses one P6 Excel progress extract into a flat list of per-activity
    rows — pure parsing, no DB access, same "fail cleanly before anything
    opens a transaction" discipline as p6_import_parse.py's own
    parse_pmxml."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Not a readable .xlsx file: {exc}") from exc

    if "TASK" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail="No \"TASK\" sheet found — not a recognisable P6 Excel export.")
    ws = wb["TASK"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise HTTPException(status_code=422, detail="TASK sheet has no header row.")

    col_index: dict[str, int] = {}
    for key, prefix in _HEADER_PREFIXES.items():
        idx = next((i for i, h in enumerate(header_row) if isinstance(h, str) and h.startswith(prefix)), None)
        if idx is None:
            raise HTTPException(status_code=422, detail=f"TASK sheet is missing an expected column starting with \"{prefix}\".")
        col_index[key] = idx

    rows: list[ParsedProgressRow] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        activity_id = raw[col_index["activity_id"]]
        if not activity_id:
            continue
        status = raw[col_index["status"]] or "Not Started"
        start = raw[col_index["start"]]
        finish = raw[col_index["finish"]]
        rows.append(ParsedProgressRow(
            activity_id=str(activity_id),
            status=str(status),
            start=start if isinstance(start, datetime) else None,
            finish=finish if isinstance(finish, datetime) else None,
            actual_cost=_decimal(raw[col_index["actual_cost"]]),
            earned_value_cost=_decimal(raw[col_index["earned_value_cost"]]),
            bac=_decimal(raw[col_index["bac"]]),
        ))
    return rows


def extract_date_from_filename(filename: str) -> date | None:
    """P6_Extract_2011-06-01.xlsx -> date(2011, 6, 1) — Maro's own extracts
    are named with the data date they represent (his own words: "in the
    names of the files, that is the data date"), so this is the intended
    way to get each extract's own as-of date rather than guessing from
    file content the flat TASK sheet doesn't actually carry (no DataDate
    column, unlike a full PMXML export)."""
    import re
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", filename)
    if m is None:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None
